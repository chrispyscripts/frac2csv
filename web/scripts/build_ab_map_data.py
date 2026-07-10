#!/usr/bin/env python3
"""Build the Alberta well-map dataset from AER's open ST37 Excel package.

Input:  ST_37_Excel/ (from https://static.aer.ca/prd/documents/sts/st37/ST_37_Excel.zip)
          ST37_BH.xlsx = UWI, TD, TVD, bottomhole lat/long, per well
          ST37_SH.xlsx = surface lat/long, per licence
Output: public/data/ab-wells.json

Alberta has ~536K wells — far too many for the browser — so we keep the
frac-relevant subset: real horizontals (surface->bottomhole displacement
above MIN_LATERAL_M). Records carry both TD (measured) and TVD (true
vertical) where available.
"""
import json
import math
import os
import sys

import openpyxl

SRC = sys.argv[1] if len(sys.argv) > 1 else \
    "/private/tmp/claude-501/-Users-chrisharder-Documents-Chris-Vault/" \
    "d6601dbf-d0d6-40de-a875-70a3852f9840/scratchpad/ST_37_Excel"
OUT = os.path.join(os.path.dirname(__file__), "..", "public", "data", "ab-wells.json")
MIN_LATERAL_M = 400.0   # keep wells with a drawable horizontal lateral


def disp_m(la1, lo1, la2, lo2):
    dy = (la2 - la1) * 111_320
    dx = (lo2 - lo1) * 111_320 * math.cos(math.radians((la1 + la2) / 2))
    return math.hypot(dx, dy)


def clean_name(s):
    return " ".join(str(s or "").split())[:60]


def main():
    sh = {}
    wb = openpyxl.load_workbook(os.path.join(SRC, "ST37_SH.xlsx"), read_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    next(it)
    for r in it:
        lic, la, lo = r[0], r[8], r[9]
        if lic and isinstance(la, (int, float)) and isinstance(lo, (int, float)):
            sh[str(lic)] = (float(la), float(lo))
    wb.close()
    print(f"surface holes: {len(sh):,}")

    out = []
    wb = openpyxl.load_workbook(os.path.join(SRC, "ST37_BH.xlsx"), read_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    next(it)
    seen = 0
    for r in it:
        seen += 1
        uwi, lic, name = r[0], r[1], r[2]
        td, tvd, bla, blo = r[5], r[6], r[13], r[14]
        if not (isinstance(bla, (int, float)) and isinstance(blo, (int, float))):
            continue
        s = sh.get(str(lic))
        if not s:
            continue
        sla, slo = s
        disp = disp_m(sla, slo, float(bla), float(blo))
        if disp < MIN_LATERAL_M:
            continue          # vertical / deviated-but-not-lateral: drop for size
        if isinstance(td, (int, float)) and td > 0 and disp > td * 1.2:
            continue          # lateral longer than the hole = bad coordinates
        rec = {"wa": str(lic), "n": clean_name(name),
               "la": round(sla, 5), "lo": round(slo, 5),
               "bla": round(float(bla), 5), "blo": round(float(blo), 5)}
        if uwi:
            rec["u"] = [str(uwi)]
        if isinstance(td, (int, float)) and td > 0:
            rec["td"] = round(float(td), 1)
        if isinstance(tvd, (int, float)) and tvd > 0:
            rec["tvd"] = round(float(tvd), 1)
        out.append(rec)
    wb.close()
    print(f"bottom holes scanned: {seen:,} -> {len(out):,} horizontals kept")

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w") as f:
        json.dump({"v": 1, "source": "Alberta Energy Regulator ST37 (open data)",
                   "note": f"horizontals only (lateral > {MIN_LATERAL_M:.0f} m)",
                   "wells": out}, f, separators=(",", ":"))
    print(f"wrote {OUT} ({os.path.getsize(OUT) // 1024} KB)")


if __name__ == "__main__":
    main()
